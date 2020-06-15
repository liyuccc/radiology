import os
import six
from radiomics import featureextractor, setVerbosity
from radiomics.featureextractor import RadiomicsFeatureExtractor
from openpyxl import Workbook
from extrat import fileHanding


def extract_save(extractor,ROIClass,Modal,name):
    dataDir = '/media/lyu/841G/DATA/Gliomas/ALL_img'
    ROIDir = '/media/lyu/841G/DATA/Gliomas/Converted_ROI'
    file = Workbook()
    row = 1

    for id,i in enumerate(ROIClass):
        # print("id",id)
        for patient in os.listdir(dataDir):
            if patient.find(i[4:7])!= -1:
                patient_path = os.path.join(dataDir,patient)
                for m in Modal:
                    for modal in os.listdir(patient_path):
                        if modal.find(m) != -1:
                            imageName = os.path.join(dataDir, patient,modal)
                            maskName = os.path.join(ROIDir, i)
                            print('Now is extracting %s`s %s_%s' %(patient,m,name))
                            try:
                                result = extractor.execute(imageName, maskName)
                                # sheets = file.sheetnames
                                if id == 0:
                                    table = file.create_sheet(m)
                                    row = 1
                                    column = 1
                                    print("create sheet:",m)
                                else :
                                    table = file.get_sheet_by_name(m)
                                    column = 1
                                for key, val in six.iteritems(result):
                                    if row == 1:
                                        table.cell(row=1, column=1, value='case')
                                        table.cell(row=2, column=1, value=os.path.join(dataDir, i))
                                        table.cell(row=1, column=column + 1, value=key)
                                        table.cell(row=2, column=column + 1, value=float(val))
                                    else:
                                        table.cell(row=row + 1, column=1, value=os.path.join(dataDir, i))
                                        table.cell(row=row + 1, column=column + 1, value=float(val))
                                    column += 1
                                    # print(column,len(result))
                                assert len(result) == column - 1

                                # print(column,len(result))

                            except:
                                print('Warning:%s`s %s_%s have problems!!!' % (patient,m, name))
        row += 1
    file.save(os.path.join('/home/lyu/PythonProjects/radiomics/Feature/Multi model','test_multi_modal_%s.xlsx'%name))
# import settings file
params = './mine.yaml'
# params = './MR.yaml'

# obtain the feature extractor class with the setting file
extractor = RadiomicsFeatureExtractor(params)
# enable or disable
extractor.addProvenance(provenance_on=False)

setVerbosity(60)

file = Workbook()
table = file.create_sheet('data')
dataDir = '/media/lyu/841G/DATA/Gliomas/ALL'
ROIDir = '/media/lyu/841G/DATA/Gliomas/Converted_ROI'

# 分四个list存不同的ROI
nangbian_mask,shizhi_mask,zhongliu_mask,shuizhong_mask,bingbian_mask = fileHanding.classifyROI(ROIDir)
# modal = ['3DT1','rCBF','ADC','Dfast','Dslow','_f']
modal = ['3DT1']
# extract_save(extractor, nangbian_mask,modal,'nangbian')
extract_save(extractor, zhongliu_mask,modal,'zhongliu')
# extract_save(extractor, shizhi_mask,modal,'shizhi')
# extract_save(extractor, shuizhong_mask,modal,'shuizhong')
# extract_save(extractor, bingbian_mask,modal,'bingbian')