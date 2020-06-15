import os
import nibabel as nib
import numpy as np
import six
from openpyxl import Workbook
from openpyxl import load_workbook


# ROI转换
def ROIhanding(ROIDir,newDir):
    roi_name_Dir = sorted([i for i in os.listdir(ROIDir)])
    for id, i in enumerate(roi_name_Dir):
        os.mkdir(os.path.join(newDir,i))
        roi = os.listdir(os.path.join(ROIDir, i))
        for j in roi:
            read_path = os.path.join(ROIDir,i,j)
            print(read_path)
            img = nib.load(read_path).get_data()
            save_path = os.path.join(newDir,i,j)
            pairimg = nib.Nifti1Pair(img, np.eye(4))
            nib.save(pairimg, save_path)


#ROI分类
def classifyROI(ROIDir):
    zhongliu_mask = []
    bingbian_mask = []
    nangbian_mask = []
    shuizhong_mask = []
    shizhi_mask = []
    roi_name_Dir = sorted([i for i in os.listdir(ROIDir)])
    for i in roi_name_Dir:
        roi = os.listdir(os.path.join(ROIDir,i))
        for j in roi:
            if j.find('zhongliu')!= -1:
                zhongliu_mask.append(os.path.join(i,j))
            elif j.find('bingbian')!= -1:
                bingbian_mask.append(os.path.join(i,j))
            elif j.find('nangbian')!= -1:
                nangbian_mask.append(os.path.join(i,j))
            elif j.find('shuizhong')!= -1:
                shuizhong_mask.append(os.path.join(i,j))
            elif j.find('shizhi')!= -1:
                shizhi_mask.append(os.path.join(i,j))
    return nangbian_mask,shizhi_mask,zhongliu_mask,shuizhong_mask,bingbian_mask


# 特征提取并保存
def extract_save(extractor,dataDir,name,Resample):
    file = Workbook()
    table = file.create_sheet('data')
    row = 1

    for id, patient in enumerate(sorted(os.listdir(dataDir))):
        column = 1
        for resample in Resample:
            imagepath = os.path.join(dataDir, patient, resample)
            # maskpath = os.path.join(dataDir, patient, resample)
            imageName = [image for image in os.listdir(imagepath) if image.endswith('_img.nii.gz')]
            maskName = [mask for mask in os.listdir(imagepath) if mask.endswith('_label.nii.gz')]
            print(imagepath)
            print('Now is extracting %s`s %s %s' %(patient,resample,name))
        # calculate the features
        #     try:
            result = extractor.execute(os.path.join(imagepath,imageName[0]), os.path.join(imagepath,maskName[0]))
            for key, val in six.iteritems(result):
                if row == 1:
                    table.cell(row=1, column=1, value='case')
                    table.cell(row=2, column=1, value=patient)
                    table.cell(row=1, column=column + 1, value=key+'_'+resample)
                    table.cell(row=2, column=column + 1, value=float(val))
                else:
                    table.cell(row=row + 1, column=1, value=patient)
                    table.cell(row=row + 1, column=column + 1, value=float(val))
                column += 1
        row += 1
        assert (column - 1) % len(result)== 0
            # except:
            #     print('Warning:%s`s %s have problems!!!'%(patient,name))

    file.save(os.path.join('/home/lyu/PythonProjects/radiomics/Feature/Feature_original','3DT1_%s.xlsx'%name))


def read_File(inputpath, sheet):
    wb = load_workbook(inputpath)
    ws = wb[sheet]
    feature = []
    for row in ws.iter_rows():
        feature.append(row)
    m = ws.max_row
    n = ws.max_column
    print(m, n)
    Feature = [["0"]*n for i in range(m)]
    for i in range(0, m):
        for j in range(0, n):
            Feature[i][j] = feature[i][j].value
    Feature = np.array(Feature)
    return Feature


def write_File(data,outputpath):
    file = Workbook()
    table = file.create_sheet('data')
    print("Feature_size:",data.shape)
    for row in range(0, np.size(data, 0)):
        for col in range(0, np.size(data, 1)):
            table.cell(row=row+1,column=col+1,value=data[row,col])
    file.save(outputpath)


def reName():
    file_path1 = '/home/lyu/PythonProjects/radiomics/Feature/Feature_original/collage_3DT1_zl_feature.xlsx'
    sheet_name1 = 'Sheet1'
    collage_feats_intra = read_File(file_path1, sheet_name1)
    file_path2 = "/home/lyu/PythonProjects/radiomics/Feature/Feature_original/3DT1_zhongliu.xlsx"
    sheet_name2 = 'data'
    patient_ID = read_File(file_path2, sheet_name2)
    collage_feats_intra = np.array(collage_feats_intra)
    collage_feats_intra = collage_feats_intra.astype(float)
    win = ['three', 'five']
    dominant_orientation = ['do1', 'do2']
    thirteen = ['entropy','energy','inertia','idm','correlation'
        ,'info1','info2','sa','sv','se','da','dv','de']
    Five = ['mean','median','std','skewness','kurtosis']
    Name = []
    Resample = ['Original', '1', '2', '3', '4', '5']
    for resample in Resample:
        for i in win:
            for j in dominant_orientation:
                for k in thirteen:
                    for t in Five:
                        name = i+'_'+j+'_'+k+'_'+t+'_'+resample
                        if len(Name)!=0:
                            Name.append(name)
                        else:
                            Name = [name]
    Name = np.expand_dims(np.array(Name),axis=0)
    collage_feats_intra = np.vstack((Name,collage_feats_intra))
    collage_feats_intra = np.hstack((np.expand_dims(patient_ID[:,0],1), collage_feats_intra))
    data_outpath = '/home/lyu/PythonProjects/radiomics/Feature/Feature_original/collage_3DT1_zl.xlsx'
    write_File(collage_feats_intra, data_outpath)


# 合并
def comp():
    file_path1 = '/home/lyu/PythonProjects/radiomics/Feature/Feature_original/collage_3DT1_zl.xlsx'
    sheet_name1 = 'data'
    collage_feats_intra = read_File(file_path1, sheet_name1)
    file_path2 = "/home/lyu/PythonProjects/radiomics/Feature/Feature_original/3DT1_zhongliu.xlsx"
    sheet_name2 = 'data'
    pR = read_File(file_path2, sheet_name2)
    collage_feats_intra = np.array(collage_feats_intra)
    collage_feats_intra = np.hstack((pR, collage_feats_intra[:,1:]))
    data_outpath = '/home/lyu/PythonProjects/radiomics/Feature/Feature_original/comp.xlsx'
    write_File(collage_feats_intra, data_outpath)


if __name__=='__main__':
    comp()