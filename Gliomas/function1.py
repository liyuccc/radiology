# 提取特征,写入,读取

import os
import nibabel as nib
import numpy as np
import six
from openpyxl import Workbook
from openpyxl import load_workbook


# ROI转换
def ROIhanding(ROIDir,newDir):
    roi_name_Dir = sorted([i for i in os.listdir(ROIDir)])
    for id, i in enumerate(roi_name_Dir):
        os.mkdir(os.path.join(newDir,i))
        roi = os.listdir(os.path.join(ROIDir, i))
        for j in roi:
            read_path = os.path.join(ROIDir,i,j)
            print(read_path)
            img = nib.load(read_path).get_data()
            save_path = os.path.join(newDir,i,j)
            pairimg = nib.Nifti1Pair(img, np.eye(4))
            nib.save(pairimg, save_path)

#ROI分类
def classifyROI(ROIDir):
    zhongliu_mask = []
    bingbian_mask = []
    nangbian_mask = []
    shuizhong_mask = []
    shizhi_mask = []
    roi_name_Dir = sorted([i for i in os.listdir(ROIDir)])
    for i in roi_name_Dir:
        roi = os.listdir(os.path.join(ROIDir,i))
        for j in roi:
            if j.find('zhongliu')!= -1:
                zhongliu_mask.append(os.path.join(i,j))
            elif j.find('bingbian')!= -1:
                bingbian_mask.append(os.path.join(i,j))
            elif j.find('nangbian')!= -1:
                nangbian_mask.append(os.path.join(i,j))
            elif j.find('shuizhong')!= -1:
                shuizhong_mask.append(os.path.join(i,j))
            elif j.find('shizhi')!= -1:
                shizhi_mask.append(os.path.join(i,j))
    return nangbian_mask,shizhi_mask,zhongliu_mask,shuizhong_mask,bingbian_mask

# 特征提取并保存
def extract_save(extractor,ROIClass,dataDir,ROIDir,name):
    file = Workbook()
    table = file.create_sheet('data')
    row = 1
    for id,i in enumerate(ROIClass):
        for j in os.listdir(dataDir):
            if j.find(i[4:7])!= -1:
                patient = j.split(".")[0]
                imageName = os.path.join(dataDir, j)
                maskName = os.path.join(ROIDir, i)
                print(imageName,maskName)
                print('Now is extracting %s`s %s' %(patient,name))
            # calculate the features
                try:
                    result = extractor.execute(imageName, maskName)
                    column = 1
                    for key, val in six.iteritems(result):
                        if row == 1:
                            table.cell(row=1, column=1, value='case')
                            table.cell(row=2, column=1, value=os.path.join(dataDir, i))
                            table.cell(row=1, column=column + 1, value=key)
                            table.cell(row=2, column=column + 1, value=float(val))
                        else:
                            table.cell(row=row + 1, column=1, value=os.path.join(dataDir, i))
                            table.cell(row=row + 1, column=column + 1, value=float(val))
                        column += 1
                    row += 1
                    assert len(result) == column - 1
                except:
                    print('Warning:%s`s %s have problems!!!'%(patient,name))

    file.save(os.path.join('Feature','%s.xlsx'%name))

def read_File(inputpath, sheet):
    wb = load_workbook(inputpath)
    ws = wb.get_sheet_by_name(sheet)
    feature = []
    for row in ws.iter_rows():
        feature.append(row)
    m = ws.max_row
    n = ws.max_column
    Feature = [["0"]*n for i in range(m)]
    for i in range(0, m):
        for j in range(0, n):
            Feature[i][j] = feature[i][j].value
    Feature = np.array(Feature)
    return Feature

def write_File(data,outputpath):
    file = Workbook()
    table = file.create_sheet('data')
    print("Feature_size:",data.shape)
    for row in range(0, np.size(data, 0)):
        for col in range(0, np.size(data, 1)):
            table.cell(row=row+1,column=col+1,value=data[row,col])
    file.save(outputpath)
