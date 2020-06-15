from sklearn.linear_model import Lasso, LogisticRegression
from sklearn.preprocessing import StandardScaler
from openpyxl import Workbook
from openpyxl import load_workbook
import numpy as np
from fileHanding import read_File, write_File
from sklearn.feature_selection import SelectFromModel


def feature_select(inputpath_Feature, outputpath, alpha):
    print(str(inputpath_Feature))
    Feature = read_File(inputpath_Feature)
    Feature1 = np.delete(Feature, 0, axis=0)
    feature_scale = np.delete(Feature1, 0, axis=1)
    wb = load_workbook("./Feature/Data.xlsx")
    ws = wb.get_sheet_by_name('valid')

    # ID
    name = [Feature[i][0] for i in range(1, np.size(Feature, 0))]
    ID = []
    for i in name:
        if i.find('roi_') != -1 and i.rfind('_') != -1:
            begin = i.find('roi_')
            end = i.rfind('_')
            id = i[begin + 4:end]
            ID.append(id)

    # 生存时间
    OS = []
    for i in range(0, len(ID)):
        for j in range(2, 57):
            q = ws.cell(row=j, column=1).value
            if q.find(ID[i]) != -1:
                os = ws.cell(row=j, column=32).value
                OS.append(os)
                break
    print("Feature_before:",feature_scale.shape)

   # Lasso
    lasso = Lasso(alpha=alpha).fit(feature_scale, OS)
    # linear = LogisticRegression(penalty='l1').fit(feature_scale, OS)
    model = SelectFromModel(lasso, prefit=True)
    feature = model.transform(feature_scale)
    id = 0
    Name = []
    Coef = []
    for coef in lasso.coef_:
        if abs(coef) >= 0.0001:
        # if coef != 0:
            Coef.append(200-abs(coef))
            Name.append(Feature[0][id+1])
        id = id +1
    FEATURE = np.c_[Coef, Name].T
    FEATURE = np.r_[FEATURE, feature]
    # 按Coef行大小排序
    print("FEATURE:",FEATURE.T[np.lexsort(FEATURE[::-1, :])].T)
    file = Workbook()
    table = file.create_sheet('data')
    print("Feature_selected:", feature.shape)
    for row in range(1, np.size(feature, 0)+2):
        for col in range(1, np.size(feature, 1)+2):
            if col == 1:
                table.cell(row=row,column=1, value=Feature[row-1][0])
            elif row==1 and col!=1:
                table.cell(row=1, column=col, value=Name[col-2])
            else:
                table.cell(row=row, column=col, value=feature[row-2, col-2])
    file.save(outputpath)

if __name__ == "__main__":
    feature_select("./Feature/Feature_scale/co_zhongliu_scale.xlsx","./Feature/Feature_selected/co_zhongliu_selected.xlsx",1)
    # feature_select("./Feature/Feature_scale/co_shuizhong_scale.xlsx","./Feature/Feature_selected/co_shuizhong_selected.xlsx", 1.5)
    # feature_select("./Feature/Feature_scale/co_shizhi_scale.xlsx","./Feature/Feature_selected/co_shizhi_selected.xlsx", 1.5)
    # feature_select("./Feature/Feature_scale/co_nangbian_scale.xlsx","./Feature/Feature_selected/co_nangbian_selected.xlsx", 1.5)
    # feature_select("./Feature/Feature_scale/co_bingbian_scale.xlsx","./Feature/Feature_selected/co_bingbian_selected.xlsx", 1.5)
    # #
    # # # abs(coef)>= 0.0001
    # feature_select("./Feature/Feature_scale/py_bingbian_scale.xlsx","./Feature/Feature_selected/py_bingbian_selected.xlsx", 1)
    # feature_select("./Feature/Feature_scale/py_nangbian_scale.xlsx","./Feature/Feature_selected/py_nangbian_selected.xlsx", 1.5)
    # feature_select("./Feature/Feature_scale/py_shizhi_scale.xlsx","./Feature/Feature_selected/py_shizhi_selected.xlsx", 1.5)
    # feature_select("./Feature/Feature_scale/py_shuizhong_scale.xlsx","./Feature/Feature_selected/py_shuizhong_selected.xlsx", 1.5)
    # feature_select("./Feature/Feature_scale/py_zhongliu_scale.xlsx","./Feature/Feature_selected/py_zhongliu_selected.xlsx", 1.5)