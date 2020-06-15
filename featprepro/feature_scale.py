from openpyxl import Workbook
from openpyxl import load_workbook
import numpy as np
from sklearn import preprocessing
from extrat.fileHanding import read_File
from collections import Counter
import math

# 读取特征
def feature_scale(inputpath,outputpath, sheet_name):
    #Feature拆分
    Feature = read_File(inputpath,sheet_name)
    # print(Feature[0,10686])
    Feature1 = np.delete(Feature, 0, axis=0)
    feature = np.delete(Feature1,0,axis=1)
    print(feature.shape)
    for j in range(0, len(Feature[0])):
        Feature[0][j] = Feature[0][j].replace('-','_')
    # 删除超过1/2相等的冗余特征
    my_list = []
    for i in range(0, feature.shape[1]):
        if None in feature[:,i] :
            my_list.append(i+1)
            continue
        maxNum_sample = Counter(feature[:,i]).most_common(1)[0][1]
        if maxNum_sample>=feature.shape[0]/2:
            my_list.append(i+1)
    Feature = np.delete(Feature, my_list, axis=1)
    Feature1 = np.delete(Feature, 0, axis=0)
    feature = np.delete(Feature1, 0, axis=1)
    # 预处理
    # Feature_scale
    min_max_scaler = preprocessing.MinMaxScaler()
    Feature_min_max = min_max_scaler.fit_transform(feature)
    feature_scaled = preprocessing.scale(Feature_min_max)

    # feature_scaled = preprocessing.normalize(feature, norm='l2',axis=0)

    print(np.size(feature_scaled, 0), np.size(feature_scaled, 1))

    WB = Workbook()
    WS_scale = WB.create_sheet("3DT1")
    for i in range(1, np.size(feature_scaled,0)+2):
        for j in range(1, np.size(feature_scaled,1)+2):
            if i == 1:
                WS_scale.cell(row=1, column=j, value=Feature[0][j-1])
            elif j == 1:
                WS_scale.cell(row=i, column=1, value=Feature[i-1][0])
            else:
                WS_scale.cell(row=i, column=j, value=feature_scaled[i-2][j-2])

    WB.save(outputpath)

if __name__ == '__main__':
    feature_scale("/home/lyu/PythonProjects/radiomics/Feature/Feature_original/3DT1_zhongliu.xlsx",
                  "/home/lyu/PythonProjects/radiomics/Feature/Feature_norm/3DT1_zhongliu_Scaled.xlsx",
                  'data')
    # feature_scale("/home/lyu/PythonProjects/radiomics/Feature/Feature_original/collage_3DT1_zl.xlsx",
    #               "/home/lyu/PythonProjects/radiomics/Feature/Feature_norm/collage_3DT1_zl_Scaled.xlsx",
    #               'data')
    # feature_scale("/home/lyu/PythonProjects/radiomics/Feature/Feature_original/comp.xlsx",
    #               "/home/lyu/PythonProjects/radiomics/Feature/Feature_norm/comp_Normed.xlsx",
    #               'data')